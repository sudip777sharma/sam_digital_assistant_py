from pynput.keyboard import Key, Controller
import wolframalpha
import datetime
import pyttsx3
import speech_recognition as sr
import wikipedia
import webbrowser
import smtplib
import openpyxl
import pygame, sys
from pygame.locals import *
from time import sleep
import os
from gtts import gTTS
from mutagen.mp3 import MP3

pygame.mixer.init()
pygame.init()
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
BLUE2 = (80, 185, 250)
BLUE3 = (0, 150, 255)
BLUE1 = (20, 90, 145)
window_x = 270
circle_x = window_x / 2
circle_y = 60
window_y = 320


engine = pyttsx3.init('sapi5')
# engine.save_to_file("the text I want to save as audio", "C:\\Users\\sudip\\PycharmProjects\\jarvis\\voice.mp3")
voices = engine.getProperty('voices')
# print(voices)
engine.setProperty('voice', voices[0].id)

def sendEmail(to, content):
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.login('sudip777sharma@gmail.com', 'JARVISSTARCK')
    speak(f"please wait sir, sending email to {to}")
    server.sendmail('sudip777sharma@gmail.com', to, content)
    server.close()


def detect_language(text):
    c = 0
    c_h = 0
    c_e = 0
    for char in text:
        maxchar = max(char)
        decision = u'\u0900' <= maxchar <= u'\u097f'
        if decision == True:
            c = 1
        if c == 1:
            if maxchar == " ":
                c_h = c_h + 1
        if decision == False:
            if maxchar != " ":
                c = -1
        if c == -1:
            if maxchar == " ":
                c_e = c_e + 1

    if c_h > c_e:
        return 'hi'
    if c_h <= c_e:
        return 'en-us'

def speak_Anim(length_textaudio):
    c = 0
    while c < int((length_textaudio / 0.42) + 1):
        # print("loop length:", int(length_textaudio / 0.42 + 2))
        c += 1
        for event in pygame.event.get():
            if event.type == QUIT:
                pygame.quit()
                sys.exit()
        windowSurface.blit(text, [75, 110])
        windowSurface.blit(text1, [63, 134])
        for i in range(1, 8, 1):
            pygame.draw.circle(windowSurface, BLUE1, (int(circle_x), circle_y), 20 + i, i)  # big circle
            pygame.draw.circle(windowSurface, BLUE2, (int(circle_x), circle_y), 21, 5)  # small
            sleep(0.03)
            pygame.display.update()

        for i in range(1, 8, 1):
            pygame.draw.circle(windowSurface, BLUE1, (int(circle_x), circle_y), 30 - i, 8 - i)  # big circle
            pygame.draw.circle(windowSurface, BLUE2, (int(circle_x), circle_y), 22, 5)  # small
            pygame.display.update()
            sleep(0.03)
            pygame.draw.circle(windowSurface, BLACK, (int(circle_x), circle_y), 33, 0)
            pygame.display.update()

def lis_recog_Anim():
    pygame.mixer.music.load("lis_recognizing.mp3")
    pygame.mixer.music.play()
    c = 0
    while c < 18:
        c += 1
        windowSurface.blit(text, [75, 110])
        windowSurface.blit(text1, [63, 134])
        pygame.draw.circle(windowSurface, BLUE1, (int(circle_x), circle_y), 30, c)  # big circle
        pygame.draw.circle(windowSurface, BLUE2, (int(circle_x), circle_y), int(30 - c / 2), 4)  # small
        sleep(0.018)
        pygame.draw.circle(windowSurface, WHITE, (int(circle_x)+1, circle_y), 13, 0)
        pygame.display.update()

def speak(audio):
    print("Sam:", audio)
    engine.say(audio)
    engine.runAndWait()

def takeCommand_online():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("listening...")
        r.adjust_for_ambient_noise(source)
        # r.pause_threshold = 1
        # r.energy_threshold = 50
        audio = r.listen(source)
        print("command:", audio)
        try:
            print("recognizing...")
            lis_recog_Anim()
            query = r.recognize_google(audio, language="en-in")
            print(f"user said: {query}\n")
        except Exception as e:
            print(e)
            print("say that again please...")
            print()
            return e
    return query

def takeCommand_offline():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("listening...")
        r.adjust_for_ambient_noise(source)
        # r.pause_threshold = 1
        # r.energy_threshold = 50
        audio = r.listen(source)
        # print("command:", audio)
        try:
            print("recognizing...")
            query = r.recognize_sphinx(audio, language="en-us")
            print(f"user said: {query}\n")
        except Exception as e:
            print(e)
            print(type(e))
            er = str(e)
            print(type(er))
            print(er)
            print("say that again please...")
            print()
            return e
    return query


def wishMe():
    hour = int(datetime.datetime.now().hour)
    if hour >= 0 and hour < 12:
        return "good morning"
    elif hour >= 12 and hour < 18:
        return "good afternoon"
    else:
        return "good evening"

def win_ini():
    pygame.draw.circle(windowSurface, BLUE1, (int(circle_x), circle_y), 30, 10)  # big circle
    pygame.draw.circle(windowSurface, BLUE2, (int(circle_x), circle_y), 24, 4)  # small
    pygame.display.update()

if __name__ == "__main__":
    speak("hello! boss")
    sleep(2)
    keyboard = Controller()
    c = 0
    query = " "
    er_online = " "
    wish_once = False
    windowSurface = pygame.display.set_mode((window_x, window_y), 0, 32)
    pygame.display.set_caption('Sam')

    basicFont = pygame.font.SysFont(None, 24)
    basicFont1 = pygame.font.SysFont(None, 20)
    text = basicFont.render('Hello!  i am Sam', True, WHITE, BLACK)
    text1 = basicFont1.render('how can i help you , sir!', True, BLUE3, BLACK)

    textRect = text.get_rect()
    textRect1 = text1.get_rect()
    windowSurface.fill(BLACK)
    windowSurface.blit(text, [75, 110])
    windowSurface.blit(text1, [63, 134])
    win_ini()
    while True:
        try:
            if not wish_once:
                wish_once = True
                pygame.mixer.music.load("lis_recognizing.mp3")
                pygame.mixer.music.play()
                sleep(1)
                speak("starting all system applications")
                speak(" installing all drivers")
                speak(" calibrating and examining all code processes")
                pygame.mixer.music.load("starting_installing_calibrating_examining.mp3")
                pygame.mixer.music.play()
                sleep(10)
                speak("all system has been started , now i am online")
                speak(wishMe())
                speak("hello i am sam , how can i help you sir")

            win_ini()
            path_conversation = "conversation.xlsx"
            workbook = openpyxl.load_workbook(path_conversation)
            sheet_conversation = workbook.active
            rows_conversation = sheet_conversation.max_row
            column_conversation = sheet_conversation.max_column

            if c == 0:
                print("here")
                query = takeCommand_online()
                er_online = str(query)
                query = er_online.lower()

            if 'recognition connection failed: [Errno 11001] getaddrinfo failed' in er_online:
                if c == 0:
                    c = 1
                    speak("you have no internet connection , offline mode is activating...")
                query = takeCommand_offline()
                er_offline = str(query)
                query = er_offline.lower()
            # logic for executing task based on query

            elif 'sleep for now' in query:
                sleep(10)

            elif 'wikipedia' in query:
                speak("searching wikipedia...")
                query = query.replace("wikipedia", "")
                results = wikipedia.summary(query, sentences=2)
                speak("According to wikipedia")
                print(results)
                speak(results)

            elif 'open youtube' in query:
                webbrowser.open("youtube.com")

            elif 'close' in query:
                keyboard.press(Key.alt)
                sleep(.3)
                keyboard.press(Key.f4)
                sleep(.3)
                keyboard.release(Key.f4)
                sleep(.3)
                keyboard.release(Key.alt)

            elif 'search' in query:
                os.startfile("C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe")
                sleep(1)
                keyboard.type(query.replace("search", ""))
                keyboard.press(Key.enter)

            elif 'open google' in query:
                os.startfile("C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe")
                # webbrowser.open("google.com")

            elif 'open stack overflow' in query:
                os.startfile("C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe")
                sleep(1)
                keyboard.type("www.stackoverflow.com")
                keyboard.press(Key.enter)
                # webbrowser.open("stack overflow.com")

            elif 'play music' in query:
                music_dir = "D:\\music"
                song = os.listdir(music_dir)
                print(song)
                os.startfile(os.path.join(music_dir, song[0]))

            elif query == "[Errno 13] Permission denied: \'conversation.xlsx\'":
                print("file isn't closed sir")
                speak("file isn't closed sir")
                workbook.close()

            elif query == "forget previous query" or query == "hey jarvis forget previous query" or query == "jarvis forget previous query":
                print("query forgot: ", sheet_conversation[f"A{rows_conversation}"].value)
                print("About query forgot:", sheet_conversation[f"I{rows_conversation}"].value)
                sheet_conversation[f"A{rows_conversation}"] = None
                sheet_conversation[f"I{rows_conversation}"] = None
                workbook.save("conversation.xlsx")
                speak("done sir")

            elif 'send email' in query:
                path = "C:\jfile\emails.xlsm"
                wb = openpyxl.load_workbook(path)
                sh = wb.active
                rows = sh.max_row
                column = sh.max_column
                speak("sir, to whom you want to send mail")
                person = takeCommand_online()
                email = ""
                for r in range(1, rows + 1):
                    if sh.cell(row=r, column=3).value in person or sh.cell(row=r, column=4).value in person:
                        email = sh.cell(row=r, column=1).value
                        person = sh.cell(row=r, column=3).value
                speak("what should i send sir.")
                nput = takeCommand_online()
                sendEmail(email, nput)
                speak(f"email has been sent to {person} ,sir.")
                print("email sent")

            q_mat = 0
            wq_mat = 0
            for r in range(1, rows_conversation + 1):
                if sheet_conversation.cell(row=r, column=1).value in query:
                    speak((sheet_conversation.cell(row=r, column=9).value).replace(".", ""))
                    q_mat = 1
                    break

            if q_mat == 0:
                client = wolframalpha.Client('E8T42Q-EKGA37XEX6')
                res = client.query(query)
                output = next(res.results).text
                wq_mat = 1
                speak(output.replace(".", ""))

            ans = ""
            if q_mat == 0 and wq_mat == 0:
                if query != "":
                    print("query:", query)
                    speak("sir i don't no about this can you please make me learn about this.")
                    sheet_conversation[f"A{rows_conversation+1}"] = query
                    print("query learn't , sir")
                    ans = takeCommand_online().lower()
                    if ans == "discard" or ans == "discard jarvis":
                        speak("ok, sir")
                        continue
                    print("ans: ", ans)
                    sheet_conversation[f"I{rows_conversation + 1}"] = ans
                    print("about query also learn't , sir")
                    workbook.save("conversation.xlsx")
                    speak("ok I learn't about that sir.")

            for event in pygame.event.get():
                print(event)
                if event.type == QUIT:
                    pygame.quit()
                    sys.exit()
        except Exception as e:
            print(e)
